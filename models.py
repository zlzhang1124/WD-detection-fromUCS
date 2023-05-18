#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Copyright (c) 2022. Institute of Health and Medical Technology, Hefei Institutes of Physical Science, CAS
# @Time     : 2022/9/5 08:44
# @Author   : ZL.Z
# @Email    : zzl1124@mail.ustc.edu.cn
# @Reference: None
# @FileName : models.py
# @Software : Python3.9; PyCharm; Ubuntu 18.04.5 LTS (GNU/Linux 5.4.0-79-generic x86_64)
# @Hardware : 2*X640-G30(XEON 6258R 2.7G); 3*NVIDIA GeForce RTX3090 (24G)
# @Version  : V1.0 - ZL.Z：2022/9/5 - 2022/9/13
#             First version.
# @License  : None
# @Brief    : 模型

from config import *
import datetime
import shutil
from matplotlib import pyplot as plt
from brokenaxes import brokenaxes
from sklearn.utils import resample
from sklearn.model_selection import train_test_split, StratifiedKFold, cross_val_score
from sklearn.metrics import confusion_matrix, roc_auc_score, accuracy_score, precision_recall_fscore_support
from tensorflow import keras
from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Input, LSTM, Dense, Dropout, Bidirectional, BatchNormalization, Activation
from attention import Attention
import scipy.stats as st
import math
from openpyxl import load_workbook


class ModelBase:
    """模型基类"""

    def __init__(self, data_file: str = "", model_name: str = "", feat_name='', seg_lv=6, shuffle=True):
        """
        初始化
        :param data_file: 数据集文件
        :param model_name: 模型名称
        :param feat_name: 待提取的特征名
        :param seg_lv: 按照音频时长比例分割特征集，1-6分别代表60秒音频的前1/6-6/6部分特征，默认6为全部时长
        :param shuffle: 是否打乱数据集，默认打乱
        """
        ft = ['comp_lld', 'w2v2', 'hubert', 'comp_lld+w2v2', 'comp_lld+hubert', 'w2v2+hubert', 'comp_lld+w2v2+hubert']
        assert feat_name in ft, f'仅接受以下特证名输入：{ft}'
        assert seg_lv in range(1, 7), f'仅接受以下分割等级输入：{list(range(1, 7))}'
        data_file = os.path.normpath(data_file)
        if data_file.endswith('pkl'):
            feat_data = pd.read_pickle(data_file)
        else:
            raise ValueError('无效数据，仅接受.pkl数据集文件')
        y_label = feat_data['label']
        if '+' in feat_name:  # 横向拼接所有特征
            _x_data_clf = []
            for i_subj in range(feat_data[feat_name.split('+')[0]].shape[0]):
                i_subj_data = feat_data[feat_name.split('+')[0]][i_subj]
                for j_ft in range(1, len(feat_name.split('+'))):
                    i_subj_data = np.hstack((i_subj_data, feat_data[feat_name.split('+')[j_ft]][i_subj]))
                _x_data_clf.append(i_subj_data)
        else:
            _x_data_clf = feat_data[feat_name]
        x_data_clf = []
        for i_subj in _x_data_clf:
            x_data_clf.append(i_subj)
        # shape=[样本数，时间序列帧数(分割后)，特征维数]
        x_data_clf = np.array(x_data_clf)[:, :math.ceil(seg_lv/6*np.array(x_data_clf).shape[1]), :]
        # print(x_data_clf.shape)
        # 分层抽样，避免数据不平衡问题
        self.train_data_clf, self.test_data_clf, train_label, test_label = \
            train_test_split(x_data_clf, y_label, random_state=rs, test_size=0.3, shuffle=shuffle, stratify=y_label)
        self.train_label, self.test_label = np.array(train_label), np.array(test_label)
        self.model_clf_save_dir = os.path.join(current_path, f'models/clf/{feat_name}')  # 保存模型路径
        if not os.path.exists(self.model_clf_save_dir):
            os.makedirs(self.model_clf_save_dir)
        self.model_name = model_name
        self.feat_name = feat_name
        self.n_folders = 10
        self.wb = load_workbook(perform_comp_f)  # 模型性能比较的EXCEL文件
        self.sheet = self.wb['Sheet1']

    def model_create(self) -> Sequential:
        """
        使用子类Keras Sequential API方式构建模型
        :return: 返回模型
        """
        pass

    def model_train_evaluate(self):
        """
        模型训练和评估
        :return: 交叉验证和测试集结果
        """
        res = {'acc_cv': [], 'f1_cv': [], 'roc_auc_cv': [], 'sen_cv': [], 'spec_cv': [],
               'acc_t': [], 'f1_t': [], 'roc_auc_t': [], 'sen_t': [], 'spec_t': []}
        skf = StratifiedKFold(n_splits=self.n_folders, shuffle=True, random_state=rs)  # 10折交叉验证，分层采样
        batch_size, epochs = 32, 50
        fold_index = 0
        for train_index, test_index in skf.split(self.train_data_clf, self.train_label):
            fold_index += 1
            print(f'------- FOLD {fold_index} / {skf.get_n_splits()} -------')
            model = self.model_create()
            # 使用回调函数保存.h5类型模型，保存权重、网络结构、参数等所有在验证集上准确率训练最好的信息,每个epoch保存一次
            save_model = os.path.join(self.model_clf_save_dir, f'model_{self.model_name}_{fold_index}.h5')
            cp_callback = tf.keras.callbacks.ModelCheckpoint(filepath=save_model, save_best_only=True,
                                                             monitor="val_accuracy", mode='max')
            es_callback = tf.keras.callbacks.EarlyStopping(monitor='loss', patience=10)
            model.fit(self.train_data_clf[train_index], self.train_label[train_index],
                      batch_size=batch_size, epochs=epochs, verbose=0,
                      validation_data=(self.train_data_clf[test_index], self.train_label[test_index]),
                      shuffle=True, callbacks=[cp_callback, es_callback])
            model_best = keras.models.load_model(save_model, custom_objects={'Attention': Attention})
            y_pred_proba = model_best(self.train_data_clf[test_index]).numpy().flatten()  # 标签为1的概率
            y_pred_label = (y_pred_proba > 0.5).astype("int32")
            acc = accuracy_score(self.train_label[test_index], y_pred_label)
            roc_auc = roc_auc_score(self.train_label[test_index], y_pred_proba)
            precision, recall, f1_score, support = \
                precision_recall_fscore_support(self.train_label[test_index], y_pred_label,
                                                average='binary', zero_division=1)  # 测试集各项指标
            sen = recall
            spec = scorer_sensitivity_specificity(self.train_label[test_index], y_pred_label)
            res['acc_cv'].append(acc)
            res['f1_cv'].append(f1_score)
            res['roc_auc_cv'].append(roc_auc)
            res['sen_cv'].append(sen)
            res['spec_cv'].append(spec)
            # 全部测试集重新训练
            es_callback = tf.keras.callbacks.EarlyStopping(monitor='val_accuracy', patience=10)
            model_best.fit(self.train_data_clf, self.train_label, validation_split=0.2,
                           batch_size=batch_size, epochs=epochs, verbose=0,
                           shuffle=True, callbacks=[cp_callback, es_callback])
            model_best_all = keras.models.load_model(save_model, custom_objects={'Attention': Attention})
            y_pred_proba_t = model_best_all(self.test_data_clf).numpy().flatten()
            y_pred_label_t = (y_pred_proba_t > 0.5).astype("int32")
            acc_t = accuracy_score(self.test_label, y_pred_label_t)
            roc_auc_t = roc_auc_score(self.test_label, y_pred_proba_t)
            precision_t, recall_t, f1_score_t, support_t = \
                precision_recall_fscore_support(self.test_label, y_pred_label_t,
                                                average='binary', zero_division=1)
            sen_t = recall_t
            spec_t = scorer_sensitivity_specificity(self.test_label, y_pred_label_t)
            res['acc_t'].append(acc_t)
            res['f1_t'].append(f1_score_t)
            res['roc_auc_t'].append(roc_auc_t)
            res['sen_t'].append(sen_t)
            res['spec_t'].append(spec_t)
            del model
            del model_best
            del model_best_all
            tf.keras.backend.clear_session()
            tf.compat.v1.reset_default_graph()
            setup_seed(rs)
        # 输出最优参数的10折交叉验证的各项指标
        print(f"CV Accuracy: {np.mean(res['acc_cv']):.4f}±{np.std(res['acc_cv']):.4f}")
        print(f"CV F1 score: {np.mean(res['f1_cv']):.4f}±{np.std(res['f1_cv']):.4f}")
        print(f"CV Sensitivity (Recall): {np.mean(res['sen_cv']):.4f}±{np.std(res['sen_cv']):.4f}")
        print(f"CV Specificity: {np.mean(res['spec_cv']):.4f}±{np.std(res['spec_cv']):.4f}")
        print(f"CV ROC-AUC: {np.mean(res['roc_auc_cv']):.4f}±{np.std(res['roc_auc_cv']):.4f}")
        print(f"Test set Accuracy: {np.mean(res['acc_t']):.4f}±{np.std(res['acc_t']):.4f}")
        print(f"Test set F1 score: {np.mean(res['f1_t']):.4f}±{np.std(res['f1_t']):.4f}")
        print(f"Test set Sensitivity (Recall): {np.mean(res['sen_t']):.4f}±{np.std(res['sen_t']):.4f}")
        print(f"Test set Specificity: {np.mean(res['spec_t']):.4f}±{np.std(res['spec_t']):.4f}")
        print(f"Test set ROC-AUC: {np.mean(res['roc_auc_t']):.4f}±{np.std(res['roc_auc_t']):.4f}")
        index = 0
        for column_cell in self.sheet.iter_rows():  # 遍历行
            index += 1
            if column_cell[0].value == self.feat_name:  # 每行的首个为对应的Features
                break
        self.sheet['C' + str(index + model_list.index(self.model_name))] = \
            f"{np.mean(res['acc_cv']):.4f}±{np.std(res['acc_cv']):.4f}"
        self.sheet['D' + str(index + model_list.index(self.model_name))] = \
            f"{np.mean(res['f1_cv']):.4f}±{np.std(res['f1_cv']):.4f}"
        self.sheet['E' + str(index + model_list.index(self.model_name))] = \
            f"{np.mean(res['sen_cv']):.4f}±{np.std(res['sen_cv']):.4f}"
        self.sheet['F' + str(index + model_list.index(self.model_name))] = \
            f"{np.mean(res['spec_cv']):.4f}±{np.std(res['spec_cv']):.4f}"
        self.sheet['G' + str(index + model_list.index(self.model_name))] = \
            f"{np.mean(res['roc_auc_cv']):.4f}±{np.std(res['roc_auc_cv']):.4f}"
        self.sheet['H' + str(index + model_list.index(self.model_name))] = \
            f"{np.mean(res['acc_t']):.4f}"
        self.sheet['I' + str(index + model_list.index(self.model_name))] = \
            f"{np.mean(res['f1_t']):.4f}"
        self.sheet['J' + str(index + model_list.index(self.model_name))] = \
            f"{np.mean(res['sen_t']):.4f}"
        self.sheet['K' + str(index + model_list.index(self.model_name))] = \
            f"{np.mean(res['spec_t']):.4f}"
        self.sheet['L' + str(index + model_list.index(self.model_name))] = \
            f"{np.mean(res['roc_auc_t']):.4f}"
        self.wb.save(perform_comp_f)  # 将结果写入模型性能比较的EXCEL文件
        return res


class BiLstmAttModel(ModelBase):
    """BiLSTM+Attention模型"""

    def __init__(self, model_name='BiLSTMAtt', **kwargs):
        """
        初始化
        :param **kwargs: 父类__init__参数
        """
        super().__init__(model_name=model_name, **kwargs)

    def model_create(self):
        """
        使用子类Keras Sequential API方式构建模型
        :return: 返回模型
        """
        model = Sequential()
        model.add(Input(shape=(self.train_data_clf.shape[1], self.train_data_clf.shape[-1])))
        model.add(Bidirectional(LSTM(128, return_sequences=True)))
        model.add(Dropout(0.3))
        model.add(Attention(32))
        model.add(Dense(16))
        model.add(BatchNormalization())
        model.add(Activation("relu"))
        model.add(Dense(1, activation="sigmoid"))
        # 编译模型：损失函数采用分类交叉熵，优化采用Adam，将识别准确率作为模型评估
        model.compile(loss="binary_crossentropy", optimizer='adam', metrics=["accuracy"])
        return model


def scorer_sensitivity_specificity(y_true, y_pred, sen_spec=False):
    """
    敏感性特异性指标
    :param y_true: 真实值
    :param y_pred: 预测概率
    :param sen_spec: True返回特异性，否则返回敏感性
    :return: 敏感性、特异性
    """
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred).ravel()
    sen = tp / (tp + fn)
    spec = tn / (tn + fp)
    if sen_spec:
        return sen
    else:
        return spec


def get_cv_res(model, features, labels, score, cv=None, n_jobs=-1):
    """
    交叉验证并获取评价指标的均值、标准差
    :param model: 模型：最优参数对应的分类或回归
    :param features: 特征
    :param labels: 标签
    :param score: 评价指标
    :param cv: 交叉验证拆分策略，默认10折交叉验证
    :param n_jobs: 并行运行的作业数
    :return: 评价指标的均值、标准差
    """
    if cv is None:
        cv = 10
    res = cross_val_score(model, features, labels, cv=cv, scoring=score, n_jobs=n_jobs)
    return f'{res.mean():.4f}', f'{res.std():.4f}'


def plot_bootstrap_ci(n_boot=2000, load_data=True):
    """
    在全部数据集上进行bootstrap试验，并获取对应的置信区间，绘制模型在各个特征集配置上的95%CI Bootstrapping结果
    :param n_boot: bootstrapping采样次数
    :param load_data: 是否加载基于之前模型采样预测后的结果直接计算CI
    :return: None
    """
    boot_accs, cis = [], []
    for _fn in feat_list:
        print(f"------- Getting {_fn} bootstrapping results -------\n")
        self = BiLstmAttModel(data_file=feat_file, model_name='BiLSTMAtt', feat_name=_fn)
        if load_data:
            acc_boot = np.load(os.path.join(self.model_clf_save_dir, f'bootstrapAcc_{self.model_name}.npy'))
        else:
            data_all = np.vstack((self.train_data_clf, self.test_data_clf))
            label_all = np.vstack((self.train_label.reshape((-1, 1)), self.test_label.reshape((-1, 1)))).ravel()
            acc_boot = []
            for i_boot in range(n_boot):  # 获取多次有放回重采样后的样本
                print(f'------- BOOT {i_boot + 1} / {n_boot} -------')
                i_boot_data, i_boot_label = resample(data_all, label_all, random_state=i_boot)
                accs = []
                for j_fold in range(1, self.n_folders + 1):  # 以每折模型测试
                    save_model = os.path.join(self.model_clf_save_dir, f'model_{self.model_name}_{j_fold}.h5')
                    model_clf = keras.models.load_model(save_model, custom_objects={'Attention': Attention})
                    y_pred_proba = model_clf(i_boot_data, training=False).numpy().flatten()
                    y_pred_label = (y_pred_proba > 0.5).astype("int32")
                    accs.append(accuracy_score(i_boot_label, y_pred_label))
                    del model_clf
                    tf.keras.backend.clear_session()
                    tf.compat.v1.reset_default_graph()
                    setup_seed(rs)
                acc_boot.append(np.mean(np.array(accs)))
            np.save(os.path.join(self.model_clf_save_dir, f'bootstrapAcc_{self.model_name}.npy'), acc_boot)
        boot_accs.append(np.mean(acc_boot))
        cis.append(list(st.norm.interval(0.95, loc=np.mean(acc_boot), scale=st.sem(acc_boot))))
    plt.figure(figsize=(5, 8), dpi=300)
    bax = brokenaxes(ylims=((0.825, 0.865), (0.95, 0.958)), hspace=.1, despine=False)
    # bax.set_title('Bootstrapping result with 95% CI', fontdict={'family': font_family, 'size': 16})
    bax.set_ylabel('Accuracy (95% CI)', fontdict={'family': font_family, 'size': 14}, labelpad=50)
    bax.axhline(0.861, c='gray', lw=1.2, ls='--', zorder=0)
    bax.errorbar(range(1, len(boot_accs) + 1), boot_accs, elinewidth=1.5,
                 yerr=abs(np.array(cis).T - boot_accs), fmt='bo', ecolor='k', capsize=7, lw=1.5, ms=5)
    bax.set_xticks(range(0, len(boot_accs) + 2), [''] + feat_figlabel + [''],
                   fontproperties=font_family, fontsize=12, rotation=30, rotation_mode='anchor', ha='right')
    bax.text(0.01, 0.861, '0.861  ', c='gray', fontdict={'family': font_family, 'size': 10}, ha='right')
    bax.set_xlim(0, len(boot_accs) + 1)
    for ba in bax.axs:
        for sp in plt.gca().spines:
            ba.spines[sp].set_color('k')
            ba.spines[sp].set_linewidth(1)
    bax.tick_params(direction='in', color='k', length=5, width=1)
    bax.tick_params('y', pad=6)
    bax.grid(False)
    fig_file = os.path.join(current_path, "results/bootstrapCIAcc.png")
    plt.savefig(fig_file, dpi=600, bbox_inches='tight', pad_inches=0.2)
    plt.savefig(fig_file.replace('.png', '.svg'), format='svg', bbox_inches='tight', pad_inches=0.2)
    plt.show()
    plt.close('all')


def plot_segment_metrics(feat_name='w2v2+hubert', load_data=True):
    """
    不同音频时长比例分割特征集后，在测试集上对应所有交叉验证模型的结果绘图呈现
    :param feat_name: 待提取的特征名
    :param load_data: 是否加载之前模型已评估好的结果直接获取指标数据
    :return: None
    """
    if load_data:
        res_all = pd.read_pickle(os.path.join(current_path, "models/clf_seg/segment_res.pkl"))
    else:
        res_all = pd.DataFrame()
        for i_seg in range(1, 7):
            self = BiLstmAttModel(data_file=feat_file, model_name='BiLSTMAtt', feat_name=feat_name, seg_lv=i_seg)
            self.model_clf_save_dir = os.path.join(current_path, f'models/clf_seg/seg{i_seg}')
            if not os.path.exists(self.model_clf_save_dir):
                os.makedirs(self.model_clf_save_dir)
            res = self.model_train_evaluate()
            res_seg = pd.DataFrame({'seg': [i_seg * 10]})
            for j_res in res:
                res_seg = pd.concat([res_seg, pd.DataFrame({j_res: [res[j_res]]})], axis=1)
            res_all = pd.concat([res_all, res_seg], ignore_index=True)
        res_all.to_pickle(os.path.join(current_path, "models/clf_seg/segment_res.pkl"))
    plt.figure(figsize=(9, 6), dpi=300, tight_layout=True)
    # plt.title(f'Accuracy over ten folders for {feat_name} model on test set', fontdict={'family': font_family, 'size': 16})
    plt.xlabel('Segment Duration (s)', fontdict={'family': font_family, 'size': 14})
    plt.ylabel('Metrics', fontdict={'family': font_family, 'size': 14})
    plt.plot(res_all['seg'].tolist(), np.mean(res_all['acc_t'].tolist(), axis=1), 'rs-', lw=1.5, ms=10, label='Acc')
    plt.plot(res_all['seg'].tolist(), np.mean(res_all['f1_t'].tolist(), axis=1), 'g^--', lw=1.5, ms=10, label='F1')
    plt.plot(res_all['seg'].tolist(), np.mean(res_all['sen_t'].tolist(), axis=1), 'cX-.', lw=1.5, ms=10, label='Sen')
    plt.plot(res_all['seg'].tolist(), np.mean(res_all['spec_t'].tolist(), axis=1), 'yD:', lw=1.5, ms=10, label='Spe')
    plt.plot(res_all['seg'].tolist(), np.mean(res_all['roc_auc_t'].tolist(), axis=1), 'b*-', lw=1.5, ms=10, label='AUC')
    plt.legend(loc="lower right", prop={'family': font_family, 'size': 12}, labelspacing=1.0)
    plt.ylim(0.65, 1.0)
    for sp in plt.gca().spines:
        plt.gca().spines[sp].set_color('k')
        plt.gca().spines[sp].set_linewidth(1)
    plt.gca().tick_params(direction='in', color='k', length=5, width=1)
    plt.grid(False)
    fig_file = os.path.join(current_path, f'results/segmentAcc.png')
    plt.savefig(fig_file, dpi=600, bbox_inches='tight', pad_inches=0.2)
    plt.savefig(fig_file.replace('.png', '.svg'), format='svg', bbox_inches='tight', pad_inches=0.2)
    plt.show()
    plt.close('all')


if __name__ == "__main__":
    start_time = datetime.datetime.now()
    print(
        f"---------- Start Time ({os.path.basename(__file__)}): {start_time.strftime('%Y-%m-%d %H:%M:%S')} ----------")
    current_path = os.path.dirname(os.path.realpath(__file__))
    feat_file = os.path.join(current_path, "data/features/features.pkl")
    res_path = os.path.join(current_path, r"results")
    perform_comp_f_raw = os.path.join(current_path, r"results/PerformanceComparison_blank.xlsx")
    perform_comp_f = os.path.join(current_path, r"results/PerformanceComparison.xlsx")
    shutil.copy(perform_comp_f_raw, perform_comp_f)

    feat_dict = {'comp_lld': 'ComParE', 'w2v2': 'w2v2', 'hubert': 'HuBERT', 'comp_lld+w2v2': 'ComParE+w2v2',
                 'comp_lld+hubert': 'ComParE+HuBERT', 'w2v2+hubert': 'w2v2+HuBERT', 'comp_lld+w2v2+hubert': 'all',}
    feat_list = list(feat_dict.keys())
    feat_figlabel = list(feat_dict.values())
    model_dict = {'BiLSTMAtt': BiLstmAttModel}
    model_list = list(model_dict.keys())
    # for ml in model_dict:
    #     for fn in feat_list:
    #         print(f"------- Running {ml} model; {fn} features -------\n")
    #         _model = model_dict[ml](data_file=feat_file, model_name=ml, feat_name=fn)
    #         _model.model_train_evaluate()
    plot_bootstrap_ci(load_data=True)
    plot_segment_metrics(load_data=True)

    end_time = datetime.datetime.now()
    print(f"---------- End Time ({os.path.basename(__file__)}): {end_time.strftime('%Y-%m-%d %H:%M:%S')} ----------")
    print(f"---------- Time Used ({os.path.basename(__file__)}): {end_time - start_time} ----------")
    with open(os.path.join(current_path, r"results/finished.txt"), "w") as ff:
        ff.write(f"------------------ Started at {start_time.strftime('%Y-%m-%d %H:%M:%S')} "
                 f"({os.path.basename(__file__)}) -------------------\r\n")
        ff.write(f"------------------ Finished at {end_time.strftime('%Y-%m-%d %H:%M:%S')} "
                 f"({os.path.basename(__file__)}) -------------------\r\n")
        ff.write(f"------------------ Time Used {end_time - start_time} "
                 f"({os.path.basename(__file__)}) -------------------\r\n")
